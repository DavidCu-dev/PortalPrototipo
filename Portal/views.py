import os
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required

# impórtacion de models y forms
from .forms import formDatosUsuario, formDeclaratoriaPropiedad, formDeclaratoriaCumplimientoAmbiental, formCartaNotificacion, formReporteVisitaTecnica, formInventarioOff
from .models import datosUsuarioM, declaratoriaPropiedadModel, declaCumpliAmbModel, cartaNotificacionModel, reporteVisitaTecnicaModel, inventarioOficina

# generacion de documentos
from io import BytesIO
from docxtpl import DocxTemplate
import openpyxl


def politicaPrivacidad(request):
    return render(request, 'PV.html')


@login_required(login_url='iniciarSesion')
def home(request):
    datos = datosUsuarioM.objects.filter(user=request.user).first()
    formDPex = declaratoriaPropiedadModel.objects.filter(user=request.user).exists()
    formCAex = declaCumpliAmbModel.objects.filter(user=request.user).exists()
    formCNex = cartaNotificacionModel.objects.filter(user=request.user).exists()
    formRVTex = reporteVisitaTecnicaModel.objects.filter(user=request.user).exists()
    formInventOf = inventarioOficina.objects.filter(user=request.user).exists()
    return render(request, 'home.html', {
        'datosUsuario': datos,
        'formDPexiste':formDPex,
        'formCAexiste':formCAex,
        'formCNexiste':formCNex,
        'formRVTexiste':formRVTex,
        'formInventarioExiste':formInventOf
    })



def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(
                    username=request.POST['username'], password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('datosUsuario')
            except IntegrityError:
                return render(request, 'signup.html', {
                    'form': UserCreationForm,
                    'error': 'El usuario ya existe'
                })
        return render(request, 'signup.html', {
            'form': UserCreationForm,
            'error': 'Las contraseñas no coinciden'
        })

@login_required(login_url='iniciarSesion')
def cerrarSesion(request):
    logout(request)
    return redirect('home')


def iniciarSesion(request):

    if request.method == 'GET':
        return render(request, 'iniciarSesion.html', {
            'form': AuthenticationForm
        })
    else:
        user = authenticate(
            request, username=request.POST['username'],  password=request.POST['password'])
        if user is None:
            return render(request, 'iniciarSesion.html', {
                'form': AuthenticationForm,
                'error': 'Usuario o contraseña incorrectos'
            })
        else:
            login(request, user)
            return redirect('home')


@login_required(login_url='iniciarSesion')
def datosUsuario(request):
    if request.method == 'GET':
        return render(request, 'datosUsuario.html', {
            'form': formDatosUsuario
        })
    else:
        form = formDatosUsuario(request.POST)
        nuevosDatos = form.save(commit=False)
        nuevosDatos.user = request.user
        nuevosDatos.save()
        return redirect('home')


@login_required(login_url='iniciarSesion')
def actualizarDatosUsuario(request):
    if request.method == 'GET':
        actualizacion = get_object_or_404(datosUsuarioM, user=request.user)
        form = formDatosUsuario(instance=actualizacion)
        return render(request, 'editDatUs.html', {
            'form': form
        })
    else:
        try:
            actualizacion = get_object_or_404(datosUsuarioM, user=request.user)
            form = formDatosUsuario(request.POST, instance=actualizacion)
            form.save()
            messages.success(request, 'Datos actualizados correctamente')
            return redirect('home')
        except ValueError:
            return render(request, 'editDatUs.html', {
                'form': form,
                'error': 'Error al actualizar los datos, por favor intente de nuevo.'
            })

# formularios 

@login_required(login_url='iniciarSesion')
def declaratoriaPropiedad(request):
    # Buscar si ya existen datos para este usuario
    datos = declaratoriaPropiedadModel.objects.filter(user=request.user).first()
    datosDPex = declaratoriaPropiedadModel.objects.filter(user=request.user).exists()

    # si se envia un formulario
    if request.method == 'POST':

        # si existen datos:
        if datos:
            # actualiza el registro existente (instance=datos) arriba
            form = formDeclaratoriaPropiedad(request.POST, instance=datos) 

        # Si no existen datos, crea uno nuevo
        else:
            form = formDeclaratoriaPropiedad(request.POST)

        # compruieba si el formulario cumple con las validaciones    
        if form.is_valid():
            # si es valido, guarda los datos
            instancia = form.save(commit=False)
            instancia.user = request.user  # Asegúrate de asignar el usuario si es necesario
            instancia.save()
            messages.success(request, 'Datos de Declaratoria de propiedad enviados correctamente')
            return redirect('home')
        else:
            messages.error(request, 'Por favor introdusca valores validos.')
            print(form.errors)
        
    # si no exsiosten datos:
    else:
        # Si ya existen datos, muestra el formulario con los datos
        if datos:
            form = formDeclaratoriaPropiedad(instance=datos)
        else:
            form = formDeclaratoriaPropiedad()

    return render(request, 'Documentos/declaratoriaPropiedad.html', {
        'form': form,
        'datosDPex': datosDPex
    })


@login_required(login_url='iniciarSesion')
def declaratoriaCumplimientoAmbiental(request):
    # Buscar si ya existen datos para este usuario
    datos = declaCumpliAmbModel.objects.filter(user=request.user).first()
    datosCAex = declaCumpliAmbModel.objects.filter(user=request.user).exists()

    # si se envia un formulario
    if request.method == 'POST':

        # si existen datos:
        if datos:
            # actualiza el registro existente (instance=datos) arriba
            form = formDeclaratoriaCumplimientoAmbiental(request.POST, instance=datos) 

        # Si no existen datos, crea uno nuevo
        else:
            form = formDeclaratoriaCumplimientoAmbiental(request.POST)

        # compruieba si el formulario cumple con las validaciones    
        
        if form.is_valid():
            # si es valido, guarda los datos
            print(request.POST)
            instancia = form.save(commit=False)
            instancia.user = request.user  # Asegúrate de asignar el usuario si es necesario
            instancia.save()
            messages.success(request, 'Datos de Declaratoria de Cumplimiento Ambiental enviados correctamente')
            return redirect('home')
        else:
            messages.error(request, 'Por favor introdusca valores validos.')
            print(form.errors)
    # si no exsiosten datos:
    else:
        # Si ya existen datos, muestra el formulario con los datos
        if datos:
            form = formDeclaratoriaCumplimientoAmbiental(instance=datos)
        else:
            form = formDeclaratoriaCumplimientoAmbiental()

    return render(request, 'Documentos/declaratoriaCA.html', {
        'form': form,
        'datosCAex':datosCAex
    })


@login_required(login_url='iniciarSesion')
def cartaNotificacion(request):
    # Buscar si ya existen datos para este usuario
    datos = cartaNotificacionModel.objects.filter(user=request.user).first()
    datosCNex = cartaNotificacionModel.objects.filter(user=request.user).exists()

    # si se envia un formulario
    if request.method == 'POST':

        # si existen datos:
        if datos:
            # actualiza el registro existente (instance=datos) arriba
            form = formCartaNotificacion(request.POST, instance=datos) 

        # Si no existen datos, crea uno nuevo
        else:
            form = formCartaNotificacion(request.POST)

        # compruieba si el formulario cumple con las validaciones    
        
        if form.is_valid():
            # si es valido, guarda los datos
            print(request.POST)
            instancia = form.save(commit=False)
            instancia.user = request.user  # Asegúrate de asignar el usuario si es necesario
            instancia.save()
            messages.success(request, 'Datos de Carta de Notificacion enviados correctamente')
            return redirect('home')
        else:
            messages.error(request, 'Por favor introdusca valores validos.')
            print(form.errors) 
    # si no exsiosten datos:
    else:
        # Si ya existen datos, muestra el formulario con los datos
        if datos:
            form = formCartaNotificacion(instance=datos)
        else:
            form = formCartaNotificacion()

    return render(request, 'Documentos/cartaNotidicacion.html', {
        'form': form,
        'datos': datos,
        'datosCNex' : datosCNex
    })


@login_required(login_url='iniciarSesion')
def reporteVisitaTecnica(request):
    # Buscar si ya existen datos para este usuario
    datos = reporteVisitaTecnicaModel.objects.filter(user=request.user).first()
    datosRVTex = reporteVisitaTecnicaModel.objects.filter(user=request.user).exists()

    # si se envia un formulario
    if request.method == 'POST':

        # si existen datos:
        if datos:
            # actualiza el registro existente (instance=datos) arriba
            form = formReporteVisitaTecnica(request.POST, instance=datos) 

        # Si no existen datos, crea uno nuevo
        else:
            form = formReporteVisitaTecnica(request.POST)

        # compruieba si el formulario cumple con las validaciones    
        
        if form.is_valid():
            # si es valido, guarda los datos
            instancia = form.save(commit=False)
            instancia.user = request.user  # Asegúrate de asignar el usuario si es necesario
            instancia.save()
            messages.success(request, 'Datos del Reporte de Visita Tecnica enviados correctamente')
            return redirect('home')
        else:
            messages.error(request, 'Por favor introdusca valores validos.')
            print(form.errors) 
    # si no exsiosten datos:
    else:
        # Si ya existen datos, muestra el formulario con los datos
        if datos:
            form = formReporteVisitaTecnica(instance=datos)
        else:
            form = formReporteVisitaTecnica()

    return render(request, 'Documentos/reporteVisitaTec.html', {
        'form': form,
        'datos': datos,
        'datosRVTex': datosRVTex
    })


@login_required(login_url='iniciarSesion')
def suplementosOficina(request):
    # Buscar si ya existen datos para este usuario
    datos = inventarioOficina.objects.filter(user=request.user).first()
    datosInventOf = inventarioOficina.objects.filter(user=request.user).exists()

    # si se envia un formulario
    if request.method == 'POST':

        # si existen datos:
        if datos:
            # actualiza el registro existente (instance=datos) arriba
            form = formInventarioOff(request.POST, instance=datos) 

        # Si no existen datos, crea uno nuevo
        else:
            form = formInventarioOff(request.POST)

        # compruieba si el formulario cumple con las validaciones    
        
        if form.is_valid():
            # si es valido, guarda los datos
            instancia = form.save(commit=False)
            instancia.user = request.user  # Asegúrate de asignar el usuario si es necesario
            instancia.save()
            messages.success(request, 'Datos de Inventario de Oficina enviados correctamente')
            return redirect('home')
        else:
            messages.error(request, 'Por favor introdusca valores validos.')
            print(form.errors) 
    # si no exsiosten datos:
    else:
        # Si ya existen datos, muestra el formulario con los datos
        if datos:
            form = formInventarioOff(instance=datos)
        else:
            form = formInventarioOff()

    return render(request, 'Documentos/inventOficina.html', {
        'form': form,
        'datos': datos,
        'datosInventOf' : datosInventOf
    })

# borrar registros
@login_required(login_url='iniciarSesion')
def borrarDP(request):
    borrar = declaratoriaPropiedadModel.objects.filter(user=request.user)
    if request.method == 'POST':
        borrar.delete()
        messages.success(request, 'Datos eliminados')
        return redirect('home')


@login_required(login_url='iniciarSesion')
def borrarDCA(request):
    borrar = declaCumpliAmbModel.objects.filter(user=request.user)
    if request.method == 'POST':
        borrar.delete()
        messages.success(request, 'Datos eliminados')
        return redirect('home')


@login_required(login_url='iniciarSesion')
def borrarCN(request):
    borrar = cartaNotificacionModel.objects.filter(user=request.user)
    if request.method == 'POST':
        borrar.delete()
        messages.success(request, 'Datos eliminados')
        return redirect('home')


@login_required(login_url='iniciarSesion')
def borrarRVT(request):
    borrar = reporteVisitaTecnicaModel.objects.filter(user=request.user)
    if request.method == 'POST':
        borrar.delete()
        messages.success(request, 'Datos eliminados')
        return redirect('home')


@login_required(login_url='iniciarSesion')
def borrarInvOf(request):
    borrar = inventarioOficina.objects.filter(user=request.user)
    if request.method == 'POST':
        borrar.delete()
        messages.success(request, 'Datos eliminados')
        return redirect('home')

# generar documentos

@login_required(login_url='iniciarSesion')
def wordDeclaratoriaPropiedad(request):
    user = request.user
    datos=get_object_or_404(declaratoriaPropiedadModel, user=user)
    template_path = os.path.join(settings.BASE_DIR, 'Portal', 'templates', 'templatesDocs', 'declaratoriaProp.docx')

    if not os.path.exists(template_path):
        return HttpResponse("La plantilla no existe en la ruta especificada.")

    context = {
        'dueñoDP': datos.dueñoDP,
        'empresaDP': datos.empresaDP,
        'rfcDP': datos.rfcDP,
        'domicilioDP': datos.domicilioDP,
        'emailDP': datos.emailDP,
        'telefonoDP': datos.telefonoDP,
        'reprelegalDP': datos.reprelegalDP,
    }

    print(context)
    doc = DocxTemplate(template_path)
    doc.render(context)

    buffer = BytesIO()  
    doc.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer,content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="declaratoria_propiedad_{datos.user}.docx"'
    return response


@login_required(login_url='iniciarSesion')
def wordDeclaratoriaCumplimientoAmbiental(request):
    user = request.user
    datos=get_object_or_404(declaCumpliAmbModel, user=user)
    template_path = os.path.join(settings.BASE_DIR, 'Portal', 'templates', 'templatesDocs', 'declaratoria de Cumplimiento Ambiental.docx')

    if not os.path.exists(template_path):
        return HttpResponse("La plantilla no existe en la ruta especificada.")
    context = {
        'empresaDCA': datos.empresaDCA,
        'domicilioDCA': datos.domicilioDCA,
        'representanteDCA': datos.representanteDCA,
        'rfcDCA': datos.rfcDCA,
        'ubicacionDCA': datos.ubicacionDCA,
        'expedienteDCA': datos.expedienteDCA,
        'responsableDCA': datos.responsableDCA,     
        'cedulaDCA': datos.cedulaDCA,
        'emailDCA': datos.emailDCA,
        'telefonoDCA': datos.telefonoDCA,
        'diaDCA': datos.diaDCA,
        'mesDCA': datos.mesDCA,
        'yearDCA': datos.yearDCA,
    }

    print(context)
    doc = DocxTemplate(template_path)
    doc.render(context)
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="declaratoria_cumplimiento_ambiental_{datos.user}.docx"'
    return response


@login_required(login_url='iniciarSesion')
def wordCartaNotificacion(request):
    user = request.user
    datos=get_object_or_404(cartaNotificacionModel, user=user)
    template_path = os.path.join(settings.BASE_DIR, 'Portal', 'templates', 'templatesDocs', 'Carta de Notificacion.docx')

    if not os.path.exists(template_path):
        return HttpResponse("La plantilla no existe en la ruta especificada.")
    context = {
        'lugarCN': datos.lugarCN,
        'diaCN': datos.diaCN,
        'mesCN': datos.mesCN,
        'anioCN': datos.anioCN,
        'nombreNotCN': datos.nombreNotCN,
        'domicilioNotCN': datos.domicilioNotCN,
        'nomRemiCN': datos.nomRemiCN,
        'cargoRemiCN': datos.cargoRemiCN,
        'depaRemiCN': datos.depaRemiCN,
        'motivoCN': datos.motivoCN,
        'objetivoCN': datos.objetivoCN,
        'plazoCN': datos.plazoCN,
        'ubiDepenCN': datos.ubiDepenCN,
        'horarioCN': datos.horarioCN,
        'telDepenCN': datos.telDepenCN,
        'emailDepenCN': datos.emailDepenCN,
    }

    print(context)
    doc = DocxTemplate(template_path)
    doc.render(context)
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="Carta-Notificacion-{datos.user}.docx"'
    return response


@login_required(login_url='iniciarSesion')
def wordReporteVisiTec(request):
    user = request.user
    datos=get_object_or_404(reporteVisitaTecnicaModel, user=user)
    template_path = os.path.join(settings.BASE_DIR, 'Portal', 'templates', 'templatesDocs', 'Reporte visita tecnica.docx')

    if not os.path.exists(template_path):
        return HttpResponse("La plantilla no existe en la ruta especificada.")
    
    context = {
        'ciudadRVT': datos.ciudadRVT ,
        'horaRVT': datos.horaRVT,
        'diaRVT': datos.diaRVT ,
        'mesRVT': datos.mesRVT ,
        'yearRVT': datos.yearRVT ,
        'inspectorRVT': datos.inspectorRVT ,
        'cargoRVT': datos.cargoRVT ,
        'departamentoRVT': datos.departamentoRVT ,
        'nomEstableRVT': datos.nomEstableRVT ,
        'direcEstableRVT': datos.direcEstableRVT ,
        'motivoVisitaRVT': datos.motivoVisitaRVT ,
        'condiciSeguRVT': datos.condiciSeguRVT ,
        'manejoResiduosRVT': datos.manejoResiduosRVT ,
        'cumpliNormasRVT': datos.cumpliNormasRVT ,
        'observacionesRVT': datos.observacionesRVT ,
        'nombreResponsRVT': datos.nombreResponsRVT ,
        'tipoIdentificacionRVT': datos.tipoIdentificacionRVT ,
        'numIdentificacionRVT': datos.numIdentificacionRVT ,
        'emailEstableciRVT': datos.emailEstableciRVT ,
        'telefonoEstableciRVT': datos.telefonoEstableciRVT ,
    }

    print(context)
    doc = DocxTemplate(template_path)
    doc.render(context)
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="Reporte visita tecnica {datos.user}.docx"'
    return response


@login_required(login_url='iniciarSesion')
def excelInventatio(request):
    datos = inventarioOficina.objects.filter(user=request.user).first()
    template_path = os.path.join(settings.BASE_DIR, 'Portal', 'templates', 'templatesDocs', 'Inventario Oficina.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    ws["D3"]=datos.escriCant
    ws["E3"]=datos.escriModelo
    ws["F3"]=datos.escriCondicion
    ws["G3"]=datos.escriUbi
    ws["H3"]=datos.escriObservaciones
    
    ws["D4"]=datos.sillaCant
    ws["E4"]=datos.sillaCant
    ws["F4"]=datos.sillaCondicion
    ws["G4"]=datos.sillaUbi
    ws["H4"]=datos.sillaObservacion

    ws["D5"]=datos.LapCant
    ws["E5"]=datos.LapModelo
    ws["F5"]=datos.LapCondicion
    ws["G5"]=datos.LapUbi
    ws["H5"]=datos.LapObservacion
    
    ws["D6"]=datos.proyecCant
    ws["E6"]=datos.proyecModelo
    ws["F6"]=datos.proyecCondicion
    ws["G6"]=datos.proyecUbi
    ws["H6"]=datos.proyecObservacion
    
    ws["D7"]=datos.impreCant
    ws["E7"]=datos.impreModelo
    ws["F7"]=datos.impreCondicion
    ws["G7"]=datos.impreUbi
    ws["H7"]=datos.impreObservacion
    
    ws["D8"]=datos.aguaCant
    ws["E8"]=datos.aguaModelo
    ws["F8"]=datos.aguaCondicion
    ws["G8"]=datos.aguaUbi
    ws["H8"]=datos.aguaObservacion

    ws["D9"]=datos.escobCant
    ws["E9"]=datos.escobModelo
    ws["F9"]=datos.escobCondicion
    ws["G9"]=datos.escobUbi
    ws["H9"]=datos.escobObservacion
    
    ws["D10"]=datos.extintCant
    ws["E10"]=datos.extintModelo
    ws["F10"]=datos.extintCondicion
    ws["G10"]=datos.extintUbi
    ws["H10"]=datos.extintObservacion
    
    ws["D11"]=datos.jabonCant
    ws["E11"]=datos.jabonModelo
    ws["F11"]=datos.jabonCondicion
    ws["G11"]=datos.jabonUbi
    ws["H11"]=datos.jabonObservacion
    
    ws["D12"]=datos.hojasCant
    ws["E12"]=datos.hojasModelo
    ws["F12"]=datos.hojasCondicion
    ws["G12"]=datos.hojasUbi
    ws["H12"]=datos.hojasObservacion

    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Respuesta HTTP para descargar el archivo
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=Inventario_Oficina_{datos.user}.xlsx'

    return response